import csv

import geocoder
import openpyxl

wb = openpyxl.load_workbook('上海市停车场数据.XLSX')
sheet = wb['Sheet1']

wb_1 = openpyxl.Workbook()
sheet_1 = wb.active
sheet_1['A1'] = '停车场名称'
sheet_1['B1'] = '地址'
sheet_1['C1'] = '经度'
sheet_1['D1'] = '纬度'

def zhuanhuan():
    # try:
        for i in range(2, sheet.max_row):

            print(i)
            B = sheet[f'B{i}']
            C = sheet[f'C{i}']

            jingwei_1 = geocoder.arcgis('上海市' + B.value)
            jingwei_2 = geocoder.arcgis('上海市' + C.value)

            if jingwei_1:
                with open('新.csv', 'a', newline='') as f:
                    writer = csv.writer(f)

                    writer.writerow([B.value, C.value, jingwei_1.latlng[1], jingwei_1.latlng[0]])
                    # print([B.value,C.value,jingwei_1.latlng[0],jingwei_1.latlng[1]])

            elif jingwei_2:
                with open('新.csv', 'a', newline='') as f:
                    writer = csv.writer(f)

                    writer.writerow([B.value, C.value, jingwei_2.latlng[1], jingwei_2.latlng[0]])
                    print([B.value, C.value, jingwei_2.latlng[0], jingwei_2.latlng[1]])

            else:
                with open('新.csv', 'a', newline='') as f:
                    writer = csv.writer(f)

                    writer.writerow([B.value, C.value, None, None])
            print(f'已转换完第{i}条地址......')
    # except:
    #     print(f'第{i}条地址未转换......')


def main():
    zhuanhuan()


if __name__ == '__main__':
    main()
